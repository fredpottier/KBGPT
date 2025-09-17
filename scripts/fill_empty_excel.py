import json
import os
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
from openai.types.chat import ChatCompletionUserMessageParam
from openpyxl import load_workbook
from qdrant_client.models import FieldCondition, Filter, MatchValue, ScoredPoint

from import_logging import setup_logging  # Ajout gestion des logs
from utils.shared_clients import (
    get_openai_client,
    get_qdrant_client,
    get_sentence_transformer,
)

# === CONFIGURATION ===
ROOT = Path(__file__).parent.parent.resolve()
LOGS_DIR = ROOT / "logs"
REJECT_LOG = LOGS_DIR / "rejected_chunks.log"
NO_MATCH_LOG = LOGS_DIR / "questions_no_match.log"
logger = setup_logging(LOGS_DIR, "fill_empty_excel_debug.log")

# Utilise la même logique que ingest_pptx_via_gpt.py pour choisir le modèle
EMB_MODEL_NAME = os.getenv("EMB_MODEL_NAME", "intfloat/multilingual-e5-base")
GPT_MODEL = os.getenv("GPT_MODEL", "gpt-4o")
COLLECTION_NAME = os.getenv("QDRANT_COLLECTION", "sap_kb")


# Custom HTTP client to ignore system envs (e.g., proxies)
model = get_sentence_transformer(EMB_MODEL_NAME)
openai_client = get_openai_client()
qdrant_client = get_qdrant_client()


def load_meta(meta_path: str | Path) -> dict[str, Any]:
    with open(meta_path, "r", encoding="utf-8") as f:
        return json.load(f)


def search_qdrant(
    question: str, solution: str, top_k: int = 5
) -> Sequence[ScoredPoint]:
    emb = model.encode([f"passage: Q: {question}"], normalize_embeddings=True)[
        0
    ].tolist()
    search_filter = Filter(
        must=[
            FieldCondition(key="main_solution", match=MatchValue(value=solution)),
            FieldCondition(key="type", match=MatchValue(value="rfp_qa")),
        ]
    )
    try:
        results = qdrant_client.search(
            collection_name=COLLECTION_NAME,
            query_vector=emb,
            limit=top_k,
            query_filter=search_filter,
        )
        logger.info(f"Recherche Qdrant pour '{question}' : {len(results)} résultats")
        return results
    except Exception as e:
        logger.warning(f"Erreur Qdrant pour '{question}': {e}")
        return []


def build_gpt_answer(question: str, context_chunks: Iterable[ScoredPoint]) -> str:
    context_parts: list[str] = []
    for chunk in context_chunks:
        payload = chunk.payload or {}
        text = payload.get("text") if isinstance(payload, dict) else ""
        context_parts.append(text if isinstance(text, str) else "")
    context = "\n\n".join(context_parts)
    prompt = (
        f"Voici une question métier SAP :\n{question}\n\n"
        f"Voici des extraits de documents pouvant aider à répondre :\n{context}\n\n"
        "Utilise uniquement ces informations pour rédiger une réponse claire et concise à la question. "
        "Réponds uniquement dans la langue de la question. "
        "Si aucune information pertinente n'est trouvée, réponds uniquement par 'Aucune réponse trouvée'."
    )
    try:
        user_message: ChatCompletionUserMessageParam = {
            "role": "user",
            "content": prompt,
        }
        response = openai_client.chat.completions.create(
            model=GPT_MODEL,
            messages=[user_message],
            temperature=0.2,
            max_tokens=512,
        )
        if not response.choices:
            raise ValueError("no completion choices returned")
        message = getattr(response.choices[0], "message", None)
        content = getattr(message, "content", None) if message else None
        if not isinstance(content, str):
            raise ValueError("completion has no textual content")
        answer = content.strip()
        logger.info(f"Réponse GPT pour '{question}': {answer[:80]}...")
        return answer
    except Exception as e:
        logger.warning(f"Erreur GPT pour '{question}': {e}")
        return "Aucune réponse trouvée"


def excel_col_letter_to_index(col_letter: str) -> int:
    col_letter = col_letter.upper()
    col_idx = 0
    for i, c in enumerate(reversed(col_letter)):
        col_idx += (ord(c) - ord("A") + 1) * (26**i)
    return col_idx - 1


def write_answers_to_excel(
    original_path: str | Path,
    df: pd.DataFrame,
    a_col_letter: str,
    a_col_idx: int,
) -> None:
    wb = load_workbook(original_path)
    ws = wb.active
    if ws is None:
        logger.error("Impossible d'accéder à la feuille active pour l'écriture.")
        return
    active_ws = ws

    for row_position, row in enumerate(df.itertuples(index=False), start=1):
        answer = row[a_col_idx]
        if pd.notna(answer):
            excel_row = row_position + 1  # +1 pour header + 1-index Excel
            active_ws.cell(
                row=excel_row,
                column=a_col_idx + 1,
                value=str(answer),
            )  # index 1-based

    output_path = (
        Path(original_path).parent / f"{Path(original_path).stem}_answered.xlsx"
    )
    wb.save(output_path)
    print(f"✅ Fichier enrichi avec styles préservés : {output_path}")
    logger.info(f"✅ Fichier enrichi avec styles préservés : {output_path}")


def filter_chunks_with_gpt(
    question: str, chunks: Iterable[ScoredPoint]
) -> list[ScoredPoint]:
    prompt_template = (
        "Question : {question}\n\n"
        "Voici un extrait de document :\n\n"
        "{chunk}\n\n"
        "Est-ce que cet extrait est pertinent pour répondre à cette question ? Réponds uniquement par : OUI ou NON."
    )

    filtered: list[ScoredPoint] = []
    rejected: list[str] = []
    for chunk in chunks:
        payload = chunk.payload or {}
        chunk_text_raw = (
            payload.get("text") if isinstance(payload, dict) else ""
        )
        chunk_text = chunk_text_raw if isinstance(chunk_text_raw, str) else ""
        prompt = prompt_template.format(question=question, chunk=chunk_text)
        try:
            user_message: ChatCompletionUserMessageParam = {
                "role": "user",
                "content": prompt,
            }
            response = openai_client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[user_message],
                temperature=0,
                max_tokens=5,
            )
            if not response.choices:
                raise ValueError("no completion choices returned")
            message = getattr(response.choices[0], "message", None)
            content = getattr(message, "content", None) if message else None
            content_lower = content.lower() if isinstance(content, str) else ""
            if "oui" in content_lower:
                filtered.append(chunk)
            else:
                rejected.append(chunk_text)
        except Exception as e:
            logger.warning(f"ChunkRAG GPT error: {e}")

    if rejected:
        with open(REJECT_LOG, "a", encoding="utf-8") as rej_log:
            for text in rejected:
                rej_log.write(
                    f"---\nQUESTION: {question}\nCHUNK NON RETENU:\n{text}\n\n"
                )

    return filtered


def clarify_question_with_gpt(question: str) -> str:
    prompt = (
        f"Voici une question métier posée dans un fichier Excel :\n\n"
        f"{question}\n\n"
        "Reformule cette question de façon claire et directe, comme si elle était posée dans un appel d'offre SAP."
        "Tu dois reformuler dans la même langue que la question d'origine."
    )
    try:
        user_message: ChatCompletionUserMessageParam = {
            "role": "user",
            "content": prompt,
        }
        response = openai_client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[user_message],
            temperature=0,
            max_tokens=100,
        )
        if not response.choices:
            raise ValueError("no completion choices returned")
        message = getattr(response.choices[0], "message", None)
        content = getattr(message, "content", None) if message else None
        if not isinstance(content, str):
            raise ValueError("clarification has no textual content")
        return content.strip()
    except Exception as e:
        logger.warning(f"Erreur clarification GPT : {e}")
        return question.strip()


def main(xlsx_path: str | Path, meta_path: str | Path) -> None:
    meta = load_meta(meta_path)
    df = pd.read_excel(
        xlsx_path, header=None
    )  # Pas de header pour éviter les problèmes de colonnes mergées
    solution_value = meta.get("solution")
    if not isinstance(solution_value, str) or not solution_value.strip():
        logger.error("Meta 'solution' manquante ou invalide.")
        return
    solution = solution_value.strip()

    logger.info(f"Traitement du fichier : {xlsx_path} ({len(df)} lignes)")

    wb = load_workbook(xlsx_path)
    visible_sheets = [
        sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == "visible"
    ]
    if not visible_sheets:
        logger.error("Aucun onglet visible dans le fichier Excel.")
        return
    ws = wb[visible_sheets[0]]

    df = pd.read_excel(xlsx_path, header=None, sheet_name=visible_sheets[0])

    merged_cells: set[str] = set()
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_cells.add(ws.cell(row=row, column=col).coordinate)

    question_col_value = meta.get("question_col")
    answer_col_value = meta.get("answer_col")
    if not isinstance(question_col_value, str) or not isinstance(
        answer_col_value, str
    ):
        logger.error("Colonnes question/réponse invalides dans le meta.")
        return

    q_idx = excel_col_letter_to_index(question_col_value)
    a_idx = excel_col_letter_to_index(answer_col_value)

    logger.debug(f"Colonnes du DataFrame : {df.columns}")
    logger.debug(f"q_idx calculé : {q_idx}")

    questions_analyzed = 0
    answers_inserted = 0

    for row_idx in range(len(df)):
        # Affiche le contenu brut de la ligne
        logger.debug(f"Ligne {row_idx} brute : {df.iloc[row_idx].to_dict()}")
        try:
            raw_question = df.iat[row_idx, q_idx]
        except Exception as e:
            logger.debug(f"⏭️ Ligne {row_idx} ignorée : erreur accès cellule ({e})")
            continue

        if pd.isna(raw_question):
            logger.debug(f"⏭️ Ligne ignorée : valeur NaN à l'index {row_idx}")
            continue

        question = str(raw_question).strip()

        if (
            pd.isna(question)
            or question == ""
            or question.lower() == "nan"
            or question.count("\n") > 3
            or len(question) < 15
            or question.lower().startswith(("section", "1.", "2.", "3."))
        ):
            logger.debug(f"⏭️ Ligne ignorée : '{question}' (non question)")
            continue

        questions_analyzed += 1
        logger.info(f"💬 Question détectée : {question}")

        clarified_question = clarify_question_with_gpt(question)
        logger.info(f"Question clarifiée : {clarified_question}")

        results = search_qdrant(clarified_question, solution)
        if not results:
            logger.info(f"Aucun chunk trouvé pour la question : {clarified_question}")
            with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                nomatch_log.write(f"{clarified_question}\n")
            continue

        filtered_chunks = filter_chunks_with_gpt(clarified_question, results)
        if not filtered_chunks:
            logger.info(
                f"Aucun chunk pertinent pour la question : {clarified_question}"
            )
            with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                nomatch_log.write(f"{clarified_question}\n")
            continue

        gpt_answer = build_gpt_answer(clarified_question, filtered_chunks)
        if gpt_answer.lower().startswith("aucune réponse trouvée"):
            logger.info(
                f"Aucune réponse pertinente pour la question : {clarified_question}"
            )
            with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                nomatch_log.write(f"{clarified_question}\n")
            continue

        wb = load_workbook(xlsx_path)
        ws = wb.active
        if ws is None:
            logger.error("Impossible d'accéder à la feuille active pour écrire la réponse.")
            continue
        active_ws = ws
        excel_row = row_idx + 1  # +1 car pas de header
        active_ws.cell(row=excel_row, column=a_idx + 1, value=gpt_answer)
        wb.save(xlsx_path)
        answers_inserted += 1
        logger.info(
            f"Réponse écrite dans Excel à la ligne {excel_row}, colonne {a_idx + 1}"
        )

        # Adaptation pour sauvegarder les réponses dans le bon fichier
        # write_answers_to_excel(xlsx_path, df, meta["answer_col"], a_idx)
    logger.info(
        f"Résumé du traitement : {questions_analyzed} questions analysées, {answers_inserted} réponses insérées"
    )


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python fillEmptyExcel.py <questions.xlsx> <meta.json>")
        logger.error("Usage: python fillEmptyExcel.py <questions.xlsx> <meta.json>")
    else:
        main(sys.argv[1], sys.argv[2])
