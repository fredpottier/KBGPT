import json
import os
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
from openai.types.chat import ChatCompletionUserMessageParam
from openpyxl import load_workbook
from qdrant_client.models import FieldCondition, Filter, MatchValue, ScoredPoint

from knowbase.common.logging import setup_logging
from knowbase.common.clients import (
    get_qdrant_client,
    get_sentence_transformer,
)
from knowbase.common.llm_router import LLMRouter, TaskType

from knowbase.config.settings import get_settings
from knowbase.config.paths import ensure_directories


class NoAnswerFoundError(Exception):
    """Exception levée quand aucune réponse pertinente n'est trouvée."""
    pass

# === CONFIGURATION ===
settings = get_settings()
LOGS_DIR = settings.logs_dir
ensure_directories([LOGS_DIR])
REJECT_LOG = LOGS_DIR / "rejected_chunks.log"
NO_MATCH_LOG = LOGS_DIR / "questions_no_match.log"
logger = setup_logging(LOGS_DIR, "fill_empty_excel_debug.log", "fill_excel_pipeline")

# Utilise la même logique que ingest_pptx_via_gpt.py pour choisir le modèle
EMB_MODEL_NAME = settings.embeddings_model
GPT_MODEL = settings.gpt_model
QA_COLLECTION_NAME = settings.qdrant_qa_collection
MAIN_COLLECTION_NAME = settings.qdrant_collection

# Seuils de confiance pour la recherche en cascade
QA_CONFIDENCE_THRESHOLD = 0.85
KB_CONFIDENCE_THRESHOLD = 0.70


# Custom HTTP client to ignore system envs (e.g., proxies)
model = get_sentence_transformer(EMB_MODEL_NAME)
llm_router = LLMRouter()
qdrant_client = get_qdrant_client()


def load_meta(meta_path: str | Path) -> dict[str, Any]:
    with open(meta_path, "r", encoding="utf-8") as f:
        return json.load(f)


def search_qa_collection(
    question: str, solution: str, top_k: int = 5
) -> Sequence[ScoredPoint]:
    """Recherche dans la collection Q/A dédiée."""
    emb = model.encode([f"passage: Q: {question}"], normalize_embeddings=True)[
        0
    ].tolist()
    search_filter = Filter(
        must=[
            FieldCondition(key="main_solution", match=MatchValue(value=solution)),
        ]
    )
    try:
        results = qdrant_client.search(
            collection_name=QA_COLLECTION_NAME,
            query_vector=emb,
            limit=top_k,
            query_filter=search_filter,
        )
        logger.info(f"Recherche Q/A pour '{question}' : {len(results)} résultats")
        return results
    except Exception as e:
        logger.warning(f"Erreur recherche Q/A pour '{question}': {e}")
        return []


def search_knowledge_base(
    question: str, solution: str, top_k: int = 10
) -> Sequence[ScoredPoint]:
    """Recherche dans la base de connaissances générale."""
    emb = model.encode([f"passage: {question}"], normalize_embeddings=True)[0].tolist()

    # Recherche avec filtres sur la solution
    search_filter = Filter(
        should=[
            # Recherche par solution principale
            FieldCondition(key="solution.main", match=MatchValue(value=solution)),
            # Recherche dans les solutions supportées
            FieldCondition(key="solution.supporting", match=MatchValue(value=solution)),
            # Recherche dans les solutions mentionnées
            FieldCondition(key="solution.mentioned", match=MatchValue(value=solution)),
        ]
    )

    try:
        results = qdrant_client.search(
            collection_name=MAIN_COLLECTION_NAME,
            query_vector=emb,
            limit=top_k,
            query_filter=search_filter,
        )
        logger.info(f"Recherche KB pour '{question}' : {len(results)} résultats")
        return results
    except Exception as e:
        logger.warning(f"Erreur recherche KB pour '{question}': {e}")
        return []


def search_hybrid(
    question: str, solution: str, top_k: int = 5, extend_search_to_kb: bool = True
) -> tuple[Sequence[ScoredPoint], str]:
    """
    Recherche en cascade : Q/A puis KB si nécessaire.

    Args:
        question: Question à rechercher
        solution: Solution SAP cible
        top_k: Nombre de résultats à retourner
        extend_search_to_kb: Si False, recherche uniquement dans Q/A collection

    Returns:
        (results, source) où source est "qa", "kb" ou "hybrid"
    """
    # 1. Recherche primaire dans les Q/A
    qa_results = search_qa_collection(question, solution, top_k)

    # Si on a des résultats Q/A avec un bon score, on les utilise
    if qa_results and qa_results[0].score >= QA_CONFIDENCE_THRESHOLD:
        logger.info(f"Réponse trouvée dans Q/A (score: {qa_results[0].score:.3f})")
        return qa_results, "qa"

    # Si l'extension à la KB est désactivée, on se contente des résultats Q/A
    if not extend_search_to_kb:
        if qa_results:
            logger.info(
                f"Recherche limitée aux Q/A - Retour résultats Q/A (meilleur score: {qa_results[0].score:.3f})"
            )
            return qa_results, "qa"
        else:
            logger.info("Recherche limitée aux Q/A - Aucun résultat trouvé")
            return [], "qa"

    # 2. Recherche dans la KB générale (uniquement si extend_search_to_kb=True)
    kb_results = search_knowledge_base(question, solution, top_k * 2)

    # 3. Décision sur la source à utiliser
    if qa_results and kb_results:
        # Comparer les meilleurs scores
        best_qa_score = qa_results[0].score if qa_results else 0
        best_kb_score = kb_results[0].score if kb_results else 0

        if (
            best_qa_score >= QA_CONFIDENCE_THRESHOLD * 0.9
        ):  # Seuil Q/A légèrement réduit
            logger.info(
                f"Utilisation Q/A malgré KB disponible (QA: {best_qa_score:.3f} vs KB: {best_kb_score:.3f})"
            )
            return qa_results, "qa"
        elif best_kb_score >= KB_CONFIDENCE_THRESHOLD:
            logger.info(f"Utilisation KB (score: {best_kb_score:.3f})")
            return kb_results[:top_k], "kb"
        else:
            # Fusion des résultats
            logger.info(
                f"Fusion Q/A + KB (QA: {best_qa_score:.3f}, KB: {best_kb_score:.3f})"
            )
            combined = list(qa_results) + list(kb_results[: top_k - len(qa_results)])
            return combined[:top_k], "hybrid"

    elif qa_results:
        logger.info(f"Utilisation Q/A seule (score: {qa_results[0].score:.3f})")
        return qa_results, "qa"
    elif kb_results:
        logger.info(f"Utilisation KB seule (score: {kb_results[0].score:.3f})")
        return kb_results[:top_k], "kb"
    else:
        logger.warning(f"Aucun résultat pour '{question}'")
        return [], "none"


def detect_language(text: str) -> str:
    """Détecte la langue principale du texte (en/fr)."""
    french_indicators = ['est', 'sont', 'avec', 'dans', 'pour', 'sur', 'vous', 'nous', 'ils', 'elles', 'cette', 'ces']
    english_indicators = ['is', 'are', 'with', 'in', 'for', 'on', 'you', 'we', 'they', 'this', 'these', 'what', 'how', 'does']

    text_lower = text.lower()
    french_count = sum(1 for indicator in french_indicators if indicator in text_lower)
    english_count = sum(1 for indicator in english_indicators if indicator in text_lower)

    return "fr" if french_count > english_count else "en"

def build_gpt_answer(question: str, context_chunks: Iterable[ScoredPoint]) -> str:
    context_parts: list[str] = []
    chunk_info_parts: list[str] = []

    for i, chunk in enumerate(context_chunks):
        payload = chunk.payload or {}
        text = payload.get("text") if isinstance(payload, dict) else ""

        # Log détaillé pour debug - source et langue apparente du chunk
        filename = payload.get("filename", "unknown") if isinstance(payload, dict) else "unknown"
        source_collection = "rfp_qa" if "rfp" in str(filename).lower() else "knowbase"
        text_preview = text[:100].replace('\n', ' ') if isinstance(text, str) else ""

        logger.debug(f"Chunk {i+1}: {source_collection} | {filename} | Score: {chunk.score:.3f} | Aperçu: {text_preview}...")

        context_parts.append(text if isinstance(text, str) else "")
        chunk_info_parts.append(f"Source: {source_collection}, Score: {chunk.score:.3f}")

    context = "\n\n".join(context_parts)

    # Log du résumé des sources pour la question
    logger.info(f"Sources pour réponse: {' | '.join(chunk_info_parts[:3])}...")

    # Détecter la langue de la question de façon plus robuste
    def detect_question_language(text: str) -> str:
        text_lower = text.lower()

        # Mots clés anglais fréquents
        english_indicators = [
            "if", "yes", "no", "which", "what", "how", "can", "do", "does", "is", "are",
            "will", "would", "should", "could", "have", "has", "the", "and", "or", "but",
            "application", "support", "provide", "system", "service", "protocol", "user"
        ]

        # Mots clés français fréquents
        french_indicators = [
            "si", "oui", "non", "quel", "quelle", "comment", "peut", "pouvez", "est", "sont",
            "sera", "serait", "devrait", "pourrait", "avoir", "avez", "le", "la", "les", "et", "ou", "mais",
            "application", "support", "fournir", "système", "service", "protocole", "utilisateur"
        ]

        english_count = sum(1 for word in english_indicators if word in text_lower)
        french_count = sum(1 for word in french_indicators if word in text_lower)

        return "English" if english_count > french_count else "French"

    question_lang = detect_question_language(question)

    # Log de la langue détectée
    logger.info(f"Question langue détectée: {question_lang} - Question: {question[:50]}...")

    # Approche en deux étapes avec message système pour forcer la langue
    system_message = (
        f"You are an RFP assistant. You MUST respond ONLY in {question_lang}. "
        f"No matter what language the source documents are in, your response must be in {question_lang}. "
        f"This is a strict requirement that cannot be overridden."
    )

    user_content = (
        f"Question ({question_lang}): {question}\n\n"
        f"Source information:\n{context}\n\n"
        f"Requirements:\n"
        f"1. Answer ONLY in {question_lang}\n"
        f"2. Use only the provided information\n"
        f"3. We are SAP - use 'SAP' instead of 'vendor/provider/supplier'\n"
        f"4. Be direct and concise\n"
        f"5. If no answer possible, respond: 'NO_ANSWER_FOUND'\n\n"
        f"Your answer in {question_lang}:"
    )

    try:
        messages = [
            {"role": "system", "content": system_message},
            {"role": "user", "content": user_content}
        ]
        content = llm_router.complete(TaskType.SHORT_ENRICHMENT, messages)
        if not isinstance(content, str):
            raise ValueError("completion has no textual content")
        answer = content.strip()

        # Vérifier si le LLM indique qu'aucune réponse n'a été trouvée
        if answer == "NO_ANSWER_FOUND":
            raise NoAnswerFoundError(f"Aucune réponse pertinente trouvée pour: {question}")

        # Validation de la langue de la réponse
        response_lang = detect_question_language(answer)
        if response_lang != question_lang:
            logger.warning(f"⚠️ Langue incorrecte détectée! Question: {question_lang}, Réponse: {response_lang}")

            # Tentative de correction avec un prompt de traduction
            translation_prompt = f"Translate this text to {question_lang}, keeping the same meaning and technical accuracy:\n\n{answer}"

            try:
                corrected_messages = [
                    {"role": "system", "content": f"You are a technical translator. Translate accurately to {question_lang}."},
                    {"role": "user", "content": translation_prompt}
                ]
                corrected_answer = llm_router.complete(TaskType.TRANSLATION, corrected_messages)
                if isinstance(corrected_answer, str):
                    answer = corrected_answer.strip()
                    logger.info(f"✅ Réponse corrigée en {question_lang}")
            except Exception as e:
                logger.warning(f"Échec de correction de langue: {e}")

        logger.info(f"Réponse GPT pour '{question}': {answer[:80]}...")
        return answer
    except NoAnswerFoundError:
        # Re-lever l'exception pour qu'elle soit gérée plus haut
        raise
    except Exception as e:
        logger.warning(f"Erreur GPT pour '{question}': {e}")
        # En cas d'erreur technique, lever aussi NoAnswerFoundError
        raise NoAnswerFoundError(f"Erreur technique lors de la génération de réponse pour: {question}")


def excel_col_letter_to_index(col_letter: str) -> int:
    col_letter = col_letter.upper()
    col_idx = 0
    for i, c in enumerate(reversed(col_letter)):
        col_idx += (ord(c) - ord("A") + 1) * (26**i)
    return col_idx - 1


def write_answers_to_excel(
    original_path: str | Path,
    df: pd.DataFrame,
    a_col_letter: str,
    a_col_idx: int,
) -> None:
    wb = load_workbook(original_path)
    ws = wb.active
    if ws is None:
        logger.error("Impossible d'accéder à la feuille active pour l'écriture.")
        return
    active_ws = ws

    for row_position, row in enumerate(df.itertuples(index=False), start=1):
        answer = row[a_col_idx]
        if pd.notna(answer):
            excel_row = row_position + 1  # +1 pour header + 1-index Excel
            active_ws.cell(
                row=excel_row,
                column=a_col_idx + 1,
                value=str(answer),
            )  # index 1-based

    output_path = (
        Path(original_path).parent / f"{Path(original_path).stem}_answered.xlsx"
    )
    wb.save(output_path)
    print(f"✅ Fichier enrichi avec styles préservés : {output_path}")
    logger.info(f"✅ Fichier enrichi avec styles préservés : {output_path}")


def filter_chunks_with_gpt(
    question: str, chunks: Iterable[ScoredPoint]
) -> list[ScoredPoint]:
    # Détecter la langue de la question
    question_lang = detect_language(question)

    if question_lang == "fr":
        prompt_template = (
            "Question : {question}\n\n"
            "Extrait de document :\n{chunk}\n\n"
            "Instructions :\n"
            "- L'extrait doit être DIRECTEMENT pertinent pour répondre à cette question spécifique\n"
            "- L'extrait doit contenir des informations factuelles liées au sujet de la question\n"
            "- Rejette les extraits trop généraux ou qui parlent d'autre chose\n"
            "- Réponds uniquement par : OUI ou NON"
        )
    else:
        prompt_template = (
            "Question: {question}\n\n"
            "Document excerpt:\n{chunk}\n\n"
            "Instructions:\n"
            "- The excerpt must be DIRECTLY relevant to answer this specific question\n"
            "- The excerpt must contain factual information related to the question topic\n"
            "- Reject excerpts that are too general or talk about something else\n"
            "- Answer only with: YES or NO"
        )

    filtered: list[ScoredPoint] = []
    rejected: list[str] = []
    for chunk in chunks:
        payload = chunk.payload or {}
        chunk_text_raw = payload.get("text") if isinstance(payload, dict) else ""
        chunk_text = chunk_text_raw if isinstance(chunk_text_raw, str) else ""
        prompt = prompt_template.format(question=question, chunk=chunk_text)
        try:
            user_message: ChatCompletionUserMessageParam = {
                "role": "user",
                "content": prompt,
            }
            content = llm_router.complete(TaskType.FAST_CLASSIFICATION, [user_message])
            content_lower = content.lower() if isinstance(content, str) else ""
            # Accepter selon la langue détectée de la question
            if question_lang == "fr":
                accept = "oui" in content_lower
            else:
                accept = "yes" in content_lower

            if accept:
                filtered.append(chunk)
            else:
                rejected.append(chunk_text)
        except Exception as e:
            logger.warning(f"ChunkRAG GPT error: {e}")

    if rejected:
        with open(REJECT_LOG, "a", encoding="utf-8") as rej_log:
            for text in rejected:
                rej_log.write(
                    f"---\nQUESTION: {question}\nCHUNK NON RETENU:\n{text}\n\n"
                )

    return filtered


def is_valid_question(question: str) -> bool:
    """Vérifie si une question est valide pour le traitement."""
    if (
        pd.isna(question)
        or question == ""
        or question.lower() == "nan"
        or question.count("\n") > 3
        or len(question) < 15
    ):
        return False

    question_lower = question.lower().strip()

    # Détecter les en-têtes de chapitre/section
    chapter_indicators = [
        "section", "chapter", "part", "module", "component",
        "technical", "system", "integration", "authentication",
        "messaging", "user", "security", "data", "configuration",
        "environment", "workstation", "support"
    ]

    # Si c'est court (< 50 chars) et contient un indicateur de chapitre sans mot interrogatif
    if len(question) < 50:
        has_chapter_indicator = any(indicator in question_lower for indicator in chapter_indicators)
        has_question_word = any(word in question_lower for word in [
            'what', 'how', 'does', 'can', 'will', 'is', 'are', 'do', 'would',
            'could', 'should', 'which', 'when', 'where', 'why', 'who'
        ])

        # Si contient un indicateur de chapitre mais pas de mot interrogatif, probablement un en-tête
        if has_chapter_indicator and not has_question_word and not question.strip().endswith('?'):
            return False

    # Détecter les numérotations de section
    if question_lower.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.", "10.")):
        return False

    # Détecter les en-têtes en MAJUSCULES courts
    if question.isupper() and len(question) < 50:
        return False

    return True


def count_valid_questions(df, q_idx: int) -> int:
    """Compte le nombre de questions valides dans le DataFrame."""
    valid_count = 0
    for row_idx in range(len(df)):
        try:
            raw_question = df.iat[row_idx, q_idx]
            if pd.isna(raw_question):
                continue
            question = str(raw_question).strip()
            if is_valid_question(question):
                valid_count += 1
        except Exception:
            continue
    return valid_count


def clarify_question_with_gpt(question: str) -> str:
    prompt = (
        f"Voici une question métier posée dans un fichier Excel :\n\n"
        f"{question}\n\n"
        "Reformule cette question de façon claire et directe, comme si elle était posée dans un appel d'offre SAP."
        "Tu dois reformuler dans la même langue que la question d'origine."
    )
    try:
        user_message: ChatCompletionUserMessageParam = {
            "role": "user",
            "content": prompt,
        }
        content = llm_router.complete(TaskType.FAST_CLASSIFICATION, [user_message])
        if not isinstance(content, str):
            raise ValueError("clarification has no textual content")
        return content.strip()
    except Exception as e:
        logger.warning(f"Erreur clarification GPT : {e}")
        return question.strip()


def main(xlsx_path: str | Path, meta_path: str | Path, progress_callback=None) -> None:
    meta = load_meta(meta_path)
    solution_value = meta.get("solution")
    if not isinstance(solution_value, str) or not solution_value.strip():
        logger.error("Meta 'solution' manquante ou invalide.")
        return
    solution = solution_value.strip()

    # Lire l'option de recherche étendue (par défaut False pour RFP)
    extend_search_to_kb = meta.get("extend_search_to_kb", False)
    if isinstance(extend_search_to_kb, str):
        extend_search_to_kb = extend_search_to_kb.lower() in ("true", "1", "yes")

    search_mode = "Q/A + KB" if extend_search_to_kb else "Q/A uniquement"
    logger.info(f"🔍 Mode de recherche configuré : {search_mode}")

    wb = load_workbook(xlsx_path)

    # Déterminer l'onglet cible : spécifié par l'utilisateur ou premier visible
    target_sheet = None
    if "sheet_name" in meta and isinstance(meta["sheet_name"], str):
        specified_sheet = meta["sheet_name"].strip()
        if specified_sheet and specified_sheet in wb.sheetnames:
            if wb[specified_sheet].sheet_state == "visible":
                target_sheet = specified_sheet
                logger.info(
                    f"📋 Utilisation de l'onglet spécifié pour RFP : {specified_sheet}"
                )
            else:
                logger.warning(
                    f"⚠️ Onglet spécifié '{specified_sheet}' n'est pas visible"
                )

    # Fallback : utiliser le premier onglet visible
    if not target_sheet:
        visible_sheets = [
            sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == "visible"
        ]
        if not visible_sheets:
            logger.error("Aucun onglet visible dans le fichier Excel.")
            return
        target_sheet = visible_sheets[0]
        logger.info(f"📋 Utilisation de l'onglet par défaut pour RFP : {target_sheet}")

    # Lire le fichier Excel avec l'onglet sélectionné
    df = pd.read_excel(xlsx_path, header=None, sheet_name=target_sheet)
    ws = wb[target_sheet]

    logger.info(
        f"Traitement du fichier : {xlsx_path} onglet '{target_sheet}' ({len(df)} lignes)"
    )

    merged_cells: set[str] = set()
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            rng.min_row,
            rng.min_col,
            rng.max_row,
            rng.max_col,
        )
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_cells.add(ws.cell(row=row, column=col).coordinate)

    question_col_value = meta.get("question_col")
    answer_col_value = meta.get("answer_col")
    if not isinstance(question_col_value, str) or not isinstance(answer_col_value, str):
        logger.error("Colonnes question/réponse invalides dans le meta.")
        return

    q_idx = excel_col_letter_to_index(question_col_value)
    a_idx = excel_col_letter_to_index(answer_col_value)

    logger.debug(f"Colonnes du DataFrame : {df.columns}")
    logger.debug(f"q_idx calculé : {q_idx}")

    # Compter d'abord le nombre total de questions valides pour une progression précise
    if progress_callback:
        progress_callback("Analyse du fichier", 0, 100, "Comptage des questions à traiter...")

    total_valid_questions = count_valid_questions(df, q_idx)
    logger.info(f"📊 {total_valid_questions} questions valides détectées sur {len(df)} lignes")

    if total_valid_questions == 0:
        logger.warning("Aucune question valide trouvée dans le fichier")
        if progress_callback:
            progress_callback("Terminé", 100, 100, "Aucune question à traiter")
        return

    questions_analyzed = 0
    answers_inserted = 0

    for row_idx in range(len(df)):
        # Affiche le contenu brut de la ligne
        logger.debug(f"Ligne {row_idx} brute : {df.iloc[row_idx].to_dict()}")
        try:
            raw_question = df.iat[row_idx, q_idx]
        except Exception as e:
            logger.debug(f"⏭️ Ligne {row_idx} ignorée : erreur accès cellule ({e})")
            continue

        if pd.isna(raw_question):
            logger.debug(f"⏭️ Ligne ignorée : valeur NaN à l'index {row_idx}")
            continue

        question = str(raw_question).strip()

        if not is_valid_question(question):
            logger.debug(f"⏭️ Ligne ignorée : '{question}' (non question)")
            continue

        questions_analyzed += 1

        # Mise à jour de la progression basée sur le nombre réel de questions
        progress_percentage = int((questions_analyzed / total_valid_questions) * 100)
        if progress_callback:
            progress_callback(
                "Traitement des questions",
                progress_percentage,
                100,
                f"Question {questions_analyzed}/{total_valid_questions}: {question[:50]}..."
            )

        logger.info(f"💬 Question {questions_analyzed}/{total_valid_questions}: {question}")

        clarified_question = clarify_question_with_gpt(question)
        logger.info(f"Question clarifiée : {clarified_question}")

        results, source = search_hybrid(
            clarified_question, solution, extend_search_to_kb=extend_search_to_kb
        )
        if not results:
            logger.info(f"Aucun chunk trouvé pour la question : {clarified_question}")
            with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                nomatch_log.write(f"{clarified_question}\n")
            continue

        logger.info(f"Source utilisée pour la réponse : {source}")

        filtered_chunks = filter_chunks_with_gpt(clarified_question, results)
        if not filtered_chunks:
            logger.info(
                f"Aucun chunk pertinent pour la question : {clarified_question}"
            )
            with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                nomatch_log.write(f"{clarified_question}\n")
            continue

        try:
            gpt_answer = build_gpt_answer(clarified_question, filtered_chunks)

            # Vérifier que la réponse n'est pas trop générique ou répétitive
            if len(gpt_answer.strip()) < 20:
                logger.info(f"Réponse trop courte ignorée pour : {clarified_question}")
                continue

        except NoAnswerFoundError as e:
            logger.info(f"Aucune réponse pertinente pour la question : {clarified_question}")
            with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                nomatch_log.write(f"{clarified_question}\n")
            continue

        wb = load_workbook(xlsx_path)
        ws = wb.active
        if ws is None:
            logger.error(
                "Impossible d'accéder à la feuille active pour écrire la réponse."
            )
            continue
        active_ws = ws
        excel_row = row_idx + 1  # +1 car pas de header
        active_ws.cell(row=excel_row, column=a_idx + 1, value=gpt_answer)
        wb.save(xlsx_path)
        answers_inserted += 1
        logger.info(
            f"Réponse écrite dans Excel à la ligne {excel_row}, colonne {a_idx + 1}"
        )

        # Adaptation pour sauvegarder les réponses dans le bon fichier
        # write_answers_to_excel(xlsx_path, df, meta["answer_col"], a_idx)
    # Progression finale
    if progress_callback:
        progress_callback(
            "Terminé",
            100,
            100,
            f"Traitement terminé: {answers_inserted} réponses générées sur {questions_analyzed} questions"
        )

    logger.info(
        f"Résumé du traitement : {questions_analyzed} questions analysées, {answers_inserted} réponses insérées"
    )


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python fillEmptyExcel.py <questions.xlsx> <meta.json>")
        logger.error("Usage: python fillEmptyExcel.py <questions.xlsx> <meta.json>")
    else:
        main(sys.argv[1], sys.argv[2])
