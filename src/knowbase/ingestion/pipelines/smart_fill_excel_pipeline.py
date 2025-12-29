"""
Pipeline intelligent pour le remplissage automatique de RFP Excel.

Nouvelle approche basée sur l'analyse LLM :
1. Extraction des lignes avec contenu (non mergées)
2. Analyse LLM batch pour filtrer/reformuler/fusionner les questions
3. Traitement optimisé avec mapping vers Excel

Améliorations par rapport à fill_excel_pipeline.py :
- Évite les erreurs sur cellules mergées
- Meilleure détection des questions vs en-têtes
- Fusion des questions multi-lignes
- Reformulation optimisée pour la recherche vectorielle
- Généralisation des noms d'entreprises (sauf SAP)
"""

import json
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from dataclasses import dataclass

import pandas as pd
from openai.types.chat import ChatCompletionUserMessageParam
from openpyxl import load_workbook
from qdrant_client.models import FieldCondition, Filter, MatchValue, ScoredPoint

from knowbase.common.logging import setup_logging
from knowbase.common.clients import (
    get_qdrant_client,
    get_sentence_transformer,
)
from knowbase.common.llm_router import LLMRouter, TaskType, get_llm_router

from knowbase.config.settings import get_settings
from knowbase.config.paths import ensure_directories


class NoAnswerFoundError(Exception):
    """Exception levée quand aucune réponse pertinente n'est trouvée."""

    pass


@dataclass
class ExtractedRow:
    """Représente une ligne extraite du fichier Excel."""

    excel_row: int  # Numéro de ligne Excel (1-based)
    question_text: str  # Texte de la question
    question_cell_merged: bool  # Si la cellule question est mergée
    answer_cell_merged: bool  # Si la cellule réponse est mergée


@dataclass
class ProcessedQuestion:
    """Représente une question après analyse LLM."""

    excel_rows: List[int]  # Lignes Excel sources (peut être plusieurs si fusion)
    original_texts: List[str]  # Textes originaux
    category: str  # "QUESTION", "HEADER", "UNCLEAR"
    processed_question: Optional[str]  # Question reformulée (si QUESTION)
    dependency_type: str  # "INDEPENDENT", "DEPENDENT"
    referenced_rows: List[int]  # Lignes référencées (si DEPENDENT)
    transformations: List[str]  # Liste des transformations appliquées
    company_replacements: Dict[str, str]  # Remplacements effectués
    confidence: float  # Confiance du LLM (0-1)


# === CONFIGURATION ===
settings = get_settings()
LOGS_DIR = settings.logs_dir
ensure_directories([LOGS_DIR])
REJECT_LOG = LOGS_DIR / "rejected_chunks.log"
NO_MATCH_LOG = LOGS_DIR / "questions_no_match.log"
logger = setup_logging(
    LOGS_DIR, "smart_fill_excel_debug.log", "smart_fill_excel_pipeline"
)

# Modèles et clients
EMB_MODEL_NAME = settings.embeddings_model
QA_COLLECTION_NAME = settings.qdrant_qa_collection
MAIN_COLLECTION_NAME = settings.qdrant_collection

# Seuils de confiance pour la recherche en cascade
QA_CONFIDENCE_THRESHOLD = 0.85
KB_CONFIDENCE_THRESHOLD = 0.70

# Entreprises à préserver (SAP et solutions SAP)
PRESERVE_COMPANIES = {
    "SAP",
    "S/4HANA",
    "SuccessFactors",
    "Ariba",
    "Concur",
    "Fieldglass",
    "SAP HANA",
    "Business One",
    "ByDesign",
    "Qualtrics",
}

# Clients globaux
model = get_sentence_transformer(EMB_MODEL_NAME)
llm_router = get_llm_router()
qdrant_client = get_qdrant_client()


def load_meta(meta_path: str | Path) -> Dict[str, Any]:
    """Charge le fichier de métadonnées."""
    with open(meta_path, "r", encoding="utf-8") as f:
        return json.load(f)


def excel_col_letter_to_index(col_letter: str) -> int:
    """Convertit une lettre de colonne Excel (A, B, AA) en index 0-based."""
    col_letter = col_letter.upper()
    col_idx = 0
    for i, c in enumerate(reversed(col_letter)):
        col_idx += (ord(c) - ord("A") + 1) * (26**i)
    return col_idx - 1


def extract_excel_rows(
    xlsx_path: str | Path, target_sheet: str, question_col_idx: int, answer_col_idx: int
) -> List[ExtractedRow]:
    """
    Extrait toutes les lignes avec du contenu dans la colonne question.
    Identifie aussi les cellules mergées pour éviter les erreurs d'écriture.
    """
    wb = load_workbook(xlsx_path)
    ws = wb[target_sheet]
    df = pd.read_excel(xlsx_path, header=None, sheet_name=target_sheet)

    # Identifier les cellules mergées
    merged_cells: set[str] = set()
    for rng in ws.merged_cells.ranges:
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                merged_cells.add(ws.cell(row=row, column=col).coordinate)

    extracted_rows = []

    for row_idx in range(len(df)):
        try:
            # Vérifier si la cellule question a du contenu
            raw_question = df.iat[row_idx, question_col_idx]
            if pd.isna(raw_question) or str(raw_question).strip() == "":
                continue

            question_text = str(raw_question).strip()
            excel_row = row_idx + 1  # Conversion en 1-based pour Excel

            # Vérifier si les cellules sont mergées
            question_cell_coord = ws.cell(
                row=excel_row, column=question_col_idx + 1
            ).coordinate
            answer_cell_coord = ws.cell(
                row=excel_row, column=answer_col_idx + 1
            ).coordinate

            question_cell_merged = question_cell_coord in merged_cells
            answer_cell_merged = answer_cell_coord in merged_cells

            extracted_row = ExtractedRow(
                excel_row=excel_row,
                question_text=question_text,
                question_cell_merged=question_cell_merged,
                answer_cell_merged=answer_cell_merged,
            )

            extracted_rows.append(extracted_row)
            logger.debug(
                f"Ligne {excel_row}: '{question_text[:50]}...' - Q_merged: {question_cell_merged}, A_merged: {answer_cell_merged}"
            )

        except Exception as e:
            logger.debug(f"Erreur extraction ligne {row_idx}: {e}")
            continue

    logger.info(
        f"Extraction terminée: {len(extracted_rows)} lignes avec contenu trouvées"
    )
    return extracted_rows


def analyze_questions_with_llm(
    extracted_rows: List[ExtractedRow],
) -> List[ProcessedQuestion]:
    """
    Analyse batch des questions avec LLM pour :
    1. Filtrer questions vs en-têtes
    2. Fusionner questions multi-lignes
    3. Reformuler pour optimiser la recherche
    4. Généraliser les noms d'entreprises (sauf SAP)
    """
    if not extracted_rows:
        return []

    # Préparer les données pour le LLM
    questions_data = []
    for i, row in enumerate(extracted_rows):
        questions_data.append(
            {
                "index": i,
                "excel_row": row.excel_row,
                "text": row.question_text,
                "merged_cells": row.question_cell_merged or row.answer_cell_merged,
            }
        )

    # Prompt pour l'analyse intelligente
    preserve_companies_str = ", ".join(PRESERVE_COMPANIES)

    prompt = f"""You are analyzing questions from an RFP Excel file for SAP solutions.

TASK: Analyze each row and return a structured analysis with intelligent question dependency handling.

COMPANIES TO PRESERVE: {preserve_companies_str}
- NEVER replace these (they refer to us as solution provider)
- Replace OTHER company names with "customer", "client", "customer organization"

INSTRUCTIONS:

1. CLASSIFY each row as:
   - "QUESTION": Real question or statement requiring validation
   - "HEADER": Section title, chapter header, separator
   - "UNCLEAR": Ambiguous cases

2. HANDLE QUESTION DEPENDENCIES:
   - Analyze each question in sequence
   - If a question references a previous question (using "If yes/no", "In case of", "For the above", etc.), mark it as DEPENDENT
   - DEPENDENT questions should be reformulated to include context from the referenced question
   - INDEPENDENT questions should be processed standalone

   Examples:
   Row 23: "Does your application require technical components on servers?" → INDEPENDENT
   Row 24: "If yes, please provide details" → DEPENDENT on Row 23

   Result:
   - Question 1 (standalone): "Does your application require technical components on customer servers?"
   - Question 2 (with context): "If your application requires technical components on customer servers, please provide details of the components and technical flows"

3. REFORMULATE questions:
   - For INDEPENDENT: Optimize for semantic search, make standalone
   - For DEPENDENT: Include full context from referenced question
   - Keep the same language as original
   - Make questions comprehensive and searchable

4. GENERALIZE company names:
   - Replace specific companies (except {preserve_companies_str}) with generic terms
   - Example: "Microsoft servers" → "customer servers"

DEPENDENCY DETECTION PATTERNS:
- "If yes/no..." → References previous question
- "In case of..." → May reference previous context
- "For the above..." → References previous question
- "Additionally..." → May be follow-up
- "Please specify..." → May be sub-question
- Question ending with "?" followed by instruction → Likely independent questions

INPUT DATA:
{json.dumps(questions_data, indent=2, ensure_ascii=False)}

OUTPUT FORMAT (valid JSON):
{{
  "processed_questions": [
    {{
      "excel_rows": [45],
      "original_texts": ["Original question text"],
      "category": "QUESTION|HEADER|UNCLEAR",
      "processed_question": "Reformulated question (null if HEADER)",
      "dependency_type": "INDEPENDENT|DEPENDENT",
      "referenced_rows": [23],
      "transformations": ["reformulated", "generalized", "contextualized"],
      "company_replacements": {{"CompanyName": "customer"}},
      "confidence": 0.95
    }}
  ]
}}

IMPORTANT:
- Process questions sequentially to detect dependencies
- For DEPENDENT questions, include full context from referenced questions
- Return ONLY valid JSON, no other text"""

    try:
        user_message: ChatCompletionUserMessageParam = {
            "role": "user",
            "content": prompt,
        }

        logger.info(f"Envoi de {len(extracted_rows)} questions au LLM pour analyse...")
        content = llm_router.complete(TaskType.RFP_QUESTION_ANALYSIS, [user_message])

        if not isinstance(content, str):
            raise ValueError("LLM response is not a string")

        # Nettoyer les balises markdown avant parsing JSON
        if content.startswith('```json'):
            content = content.replace('```json\n', '').replace('\n```', '')
        elif content.startswith('```'):
            content = content.replace('```\n', '').replace('\n```', '')

        # Vérifier que le JSON est complet avant parsing
        content = content.strip()
        if not content.endswith('}'):
            logger.warning(f"JSON tronqué détecté. Longueur: {len(content)} caractères")
            logger.debug(f"JSON se termine par: {content[-50:]}")
            raise ValueError("Réponse LLM tronquée - JSON incomplet")

        # Parser la réponse JSON
        response_data = json.loads(content)
        processed_questions = []

        for item in response_data.get("processed_questions", []):
            processed_question = ProcessedQuestion(
                excel_rows=item.get("excel_rows", []),
                original_texts=item.get("original_texts", []),
                category=item.get("category", "UNCLEAR"),
                processed_question=item.get("processed_question"),
                dependency_type=item.get("dependency_type", "INDEPENDENT"),
                referenced_rows=item.get("referenced_rows", []),
                transformations=item.get("transformations", []),
                company_replacements=item.get("company_replacements", {}),
                confidence=item.get("confidence", 0.0),
            )
            processed_questions.append(processed_question)

        # Statistiques et logs détaillés
        questions_count = sum(
            1 for q in processed_questions if q.category == "QUESTION"
        )
        headers_count = sum(1 for q in processed_questions if q.category == "HEADER")
        dependent_count = sum(
            1 for q in processed_questions if q.dependency_type == "DEPENDENT"
        )
        independent_count = sum(
            1 for q in processed_questions if q.dependency_type == "INDEPENDENT"
        )

        logger.info(
            f"Analyse LLM terminée: {questions_count} questions, {headers_count} en-têtes identifiés"
        )
        logger.info(
            f"Dépendances: {independent_count} indépendantes, {dependent_count} dépendantes"
        )

        # Log des dépendances détectées
        for q in processed_questions:
            if q.dependency_type == "DEPENDENT" and q.referenced_rows:
                logger.info(
                    f"Question dépendante ligne {q.excel_rows[0]} → référence lignes {q.referenced_rows}"
                )
                logger.debug(f"  Original: {q.original_texts[0][:60]}...")
                logger.debug(
                    f"  Reformulée: {q.processed_question[:80]}..."
                    if q.processed_question
                    else ""
                )

        return processed_questions

    except Exception as e:
        logger.error(f"Erreur analyse LLM: {e}")
        # Fallback: traiter toutes les lignes comme des questions
        fallback_questions = []
        for row in extracted_rows:
            if not (row.question_cell_merged or row.answer_cell_merged):
                fallback_question = ProcessedQuestion(
                    excel_rows=[row.excel_row],
                    original_texts=[row.question_text],
                    category="QUESTION",
                    processed_question=row.question_text,
                    dependency_type="INDEPENDENT",
                    referenced_rows=[],
                    transformations=["fallback"],
                    company_replacements={},
                    confidence=0.5,
                )
                fallback_questions.append(fallback_question)

        logger.warning(
            f"Fallback activé: {len(fallback_questions)} questions traitées sans analyse LLM"
        )
        return fallback_questions


def search_and_answer_questions(
    processed_questions: List[ProcessedQuestion],
    solution: str,
    extend_search_to_kb: bool = False,
) -> Dict[int, str]:
    """
    Recherche et génère les réponses pour les questions validées.
    Retourne un mapping excel_row -> answer pour les lignes à remplir.
    """
    # Importer les fonctions de recherche du pipeline original
    from knowbase.ingestion.pipelines.fill_excel_pipeline import (
        search_hybrid,
        filter_chunks_with_gpt,
        build_gpt_answer,
        NoAnswerFoundError,
    )

    answers = {}
    valid_questions = [
        q
        for q in processed_questions
        if q.category == "QUESTION" and q.processed_question
    ]

    logger.info(f"Traitement de {len(valid_questions)} questions validées...")

    for i, question_data in enumerate(valid_questions):
        question = question_data.processed_question
        excel_row = question_data.excel_rows[
            0
        ]  # Prendre la première ligne pour l'écriture

        logger.info(
            f"💬 Question {i+1}/{len(valid_questions)} (ligne {excel_row}): {question[:80]}..."
        )

        try:
            # Recherche vectorielle
            results, source = search_hybrid(
                question, solution, extend_search_to_kb=extend_search_to_kb
            )

            if not results:
                logger.info(f"Aucun chunk trouvé pour: {question}")
                with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                    nomatch_log.write(f"{question}\n")
                continue

            # Filtrage avec LLM
            filtered_chunks = filter_chunks_with_gpt(question, results)
            if not filtered_chunks:
                logger.info(f"Aucun chunk pertinent après filtrage pour: {question}")
                continue

            # Génération de la réponse
            answer = build_gpt_answer(question, filtered_chunks)

            if len(answer.strip()) < 20:
                logger.info(f"Réponse trop courte ignorée pour: {question}")
                continue

            answers[excel_row] = answer
            logger.info(f"✅ Réponse générée pour ligne {excel_row}")

        except NoAnswerFoundError as e:
            logger.info(f"Aucune réponse pertinente pour: {question}")
            with open(NO_MATCH_LOG, "a", encoding="utf-8") as nomatch_log:
                nomatch_log.write(f"{question}\n")
            continue
        except Exception as e:
            logger.error(f"Erreur traitement question ligne {excel_row}: {e}")
            continue

    logger.info(f"Génération terminée: {len(answers)} réponses créées")
    return answers


def write_answers_to_excel(
    xlsx_path: str | Path,
    target_sheet: str,
    answers: Dict[int, str],
    answer_col_idx: int,
) -> None:
    """
    Écrit les réponses dans le fichier Excel aux lignes spécifiées.
    """
    wb = load_workbook(xlsx_path)
    ws = wb[target_sheet]

    logger.info(f"Écriture de {len(answers)} réponses dans Excel...")

    for excel_row, answer in answers.items():
        try:
            ws.cell(row=excel_row, column=answer_col_idx + 1, value=answer)
            logger.debug(
                f"Réponse écrite ligne {excel_row}, colonne {answer_col_idx + 1}"
            )
        except Exception as e:
            logger.error(f"Erreur écriture ligne {excel_row}: {e}")
            continue

    wb.save(xlsx_path)
    logger.info(f"Fichier Excel sauvegardé: {xlsx_path}")


def main(xlsx_path: str | Path, meta_path: str | Path, progress_callback=None) -> None:
    """Point d'entrée principal du pipeline intelligent."""

    meta = load_meta(meta_path)
    solution = meta.get("solution", "").strip()
    if not solution:
        logger.error("Meta 'solution' manquante ou invalide.")
        return

    extend_search_to_kb = meta.get("extend_search_to_kb", False)
    if isinstance(extend_search_to_kb, str):
        extend_search_to_kb = extend_search_to_kb.lower() in ("true", "1", "yes")

    logger.info(f"🚀 Démarrage pipeline intelligent - Solution: {solution}")
    logger.info(
        f"🔍 Mode recherche: {'Q/A + KB' if extend_search_to_kb else 'Q/A uniquement'}"
    )

    # Progression
    if progress_callback:
        progress_callback(
            "Analyse du fichier", 10, 100, "Chargement du fichier Excel..."
        )

    # Configuration des colonnes
    wb = load_workbook(xlsx_path)
    target_sheet = None

    # Déterminer l'onglet cible
    if "sheet_name" in meta and isinstance(meta["sheet_name"], str):
        specified_sheet = meta["sheet_name"].strip()
        if specified_sheet and specified_sheet in wb.sheetnames:
            if wb[specified_sheet].sheet_state == "visible":
                target_sheet = specified_sheet
                logger.info(f"📋 Onglet spécifié: {specified_sheet}")

    if not target_sheet:
        visible_sheets = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
        if not visible_sheets:
            logger.error("Aucun onglet visible dans le fichier Excel.")
            return
        target_sheet = visible_sheets[0]
        logger.info(f"📋 Onglet par défaut: {target_sheet}")

    question_col_idx = excel_col_letter_to_index(meta["question_col"])
    answer_col_idx = excel_col_letter_to_index(meta["answer_col"])

    # Étape 1: Extraction des lignes
    if progress_callback:
        progress_callback(
            "Extraction", 20, 100, "Extraction des lignes avec contenu..."
        )

    extracted_rows = extract_excel_rows(
        xlsx_path, target_sheet, question_col_idx, answer_col_idx
    )

    if not extracted_rows:
        logger.warning("Aucune ligne avec contenu trouvée")
        if progress_callback:
            progress_callback("Terminé", 100, 100, "Aucune ligne à traiter")
        return

    # Étape 2: Analyse LLM
    if progress_callback:
        progress_callback(
            "Analyse IA",
            40,
            100,
            f"Analyse intelligente de {len(extracted_rows)} lignes...",
        )

    processed_questions = analyze_questions_with_llm(extracted_rows)

    # Étape 3: Recherche et réponses
    if progress_callback:
        progress_callback(
            "Génération réponses", 60, 100, "Recherche et génération des réponses..."
        )

    answers = search_and_answer_questions(
        processed_questions, solution, extend_search_to_kb
    )

    # Étape 4: Écriture Excel
    if progress_callback:
        progress_callback(
            "Sauvegarde", 90, 100, f"Écriture de {len(answers)} réponses..."
        )

    write_answers_to_excel(xlsx_path, target_sheet, answers, answer_col_idx)

    # Finalisation
    if progress_callback:
        progress_callback(
            "Terminé", 100, 100, f"Pipeline terminé: {len(answers)} réponses générées"
        )

    # Statistiques finales
    questions_count = sum(1 for q in processed_questions if q.category == "QUESTION")
    headers_count = sum(1 for q in processed_questions if q.category == "HEADER")

    logger.info(f"🎉 Pipeline intelligent terminé:")
    logger.info(f"  - {len(extracted_rows)} lignes extraites")
    logger.info(f"  - {questions_count} questions identifiées")
    logger.info(f"  - {headers_count} en-têtes filtrés")
    logger.info(f"  - {len(answers)} réponses générées")


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python smart_fill_excel_pipeline.py <questions.xlsx> <meta.json>")
        logger.error(
            "Usage: python smart_fill_excel_pipeline.py <questions.xlsx> <meta.json>"
        )
    else:
        main(sys.argv[1], sys.argv[2])
